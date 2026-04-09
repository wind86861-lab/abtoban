import io
from datetime import datetime

from aiogram import F, Router
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.bot.filters import RoleFilter
from app.db.models import ADMIN_ROLES, ORDER_STATUS_LABELS
from app.services.expense_service import EXPENSE_LABELS
from app.services.report_service import ReportService

router = Router()

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="D6E4F0")


def _style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


# ── Export menu ────────────────────────────────────────────────────────────────

@router.message(F.text == "📥 Eksport", RoleFilter(*ADMIN_ROLES))
async def export_menu(message: Message) -> None:
    builder = InlineKeyboardBuilder()
    builder.button(text="📋 Zakazlar (bu oy)", callback_data="export_orders:month")
    builder.button(text="📋 Zakazlar (hammasi)", callback_data="export_orders:all")
    builder.button(text="💸 Xarajatlar (bu oy)", callback_data="export_expenses:month")
    builder.button(text="📦 Material so'rovlar (bu oy)", callback_data="export_materials:month")
    builder.adjust(1)
    await message.answer(
        "📥 <b>Excel eksport</b>\n\nQaysi ma'lumotni yuklab olmoqchisiz?",
        reply_markup=builder.as_markup(),
    )


# ── Orders export ──────────────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("export_orders:"), RoleFilter(*ADMIN_ROLES))
async def export_orders(callback: CallbackQuery, session) -> None:
    await callback.answer("⏳ Fayl tayyorlanmoqda...")
    period = callback.data.split(":")[1]
    start, end = ReportService.period_bounds(period)

    svc = ReportService(session)
    orders = await svc.get_orders_for_export(start, end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Zakazlar"
    ws.row_dimensions[1].height = 30

    headers = [
        "№", "Zakaz raqami", "Klient", "Telefon", "Manzil",
        "Maydon (m²)", "Asfalt turi", "Jami summa", "To'langan",
        "Qarz", "Holat", "Master", "Usta", "Ish sanasi", "Yaratildi",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, o in enumerate(orders, 1):
        status_label = ORDER_STATUS_LABELS.get(o.status, o.status.value)
        asphalt = o.asphalt_type.name if o.asphalt_type else ""
        master = o.master.full_name if o.master else ""
        usta = o.usta.full_name if o.usta else ""
        work_date = o.work_date.strftime("%d.%m.%Y") if o.work_date else ""
        row = [
            i,
            o.order_number,
            o.client_name,
            o.client_phone,
            o.address or "",
            float(o.area_m2) if o.area_m2 else "",
            asphalt,
            float(o.total_price) if o.total_price else "",
            float(o.advance_paid) if o.advance_paid else 0,
            float(o.debt) if o.debt else 0,
            status_label,
            master,
            usta,
            work_date,
            o.created_at.strftime("%d.%m.%Y %H:%M"),
        ]
        ws.append(row)
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=col).fill = _ALT_FILL

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"zakazlar_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename=fname),
        caption=f"📋 Zakazlar — {len(orders)} ta yozuv",
    )


# ── Expenses export ────────────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("export_expenses:"), RoleFilter(*ADMIN_ROLES))
async def export_expenses(callback: CallbackQuery, session) -> None:
    await callback.answer("⏳ Fayl tayyorlanmoqda...")
    period = callback.data.split(":")[1]
    start, end = ReportService.period_bounds(period)

    svc = ReportService(session)
    expenses = await svc.get_expenses_for_export(start, end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Xarajatlar"

    headers = ["№", "Zakaz raqami", "Tur", "Summa", "Izoh", "Sana"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, e in enumerate(expenses, 1):
        order_num = e.order.order_number if e.order else ""
        label = EXPENSE_LABELS.get(e.expense_type, e.expense_type.value)
        ws.append([
            i, order_num, label, float(e.amount),
            e.description or "",
            e.created_at.strftime("%d.%m.%Y %H:%M"),
        ])
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=col).fill = _ALT_FILL

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"xarajatlar_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename=fname),
        caption=f"💸 Xarajatlar — {len(expenses)} ta yozuv",
    )


# ── Material requests export ───────────────────────────────────────────────────

@router.callback_query(F.data.startswith("export_materials:"), RoleFilter(*ADMIN_ROLES))
async def export_materials(callback: CallbackQuery, session) -> None:
    await callback.answer("⏳ Fayl tayyorlanmoqda...")
    period = callback.data.split(":")[1]
    start, end = ReportService.period_bounds(period)

    svc = ReportService(session)
    requests = await svc.get_material_requests_for_export(start, end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Material so'rovlar"

    headers = [
        "№", "Zakaz raqami", "Usta", "Miqdor (t)",
        "Material narx", "Dostavka", "Qo'shimcha", "Jami", "Holat", "Sana",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    STATUS_MAP = {"pending": "Kutilmoqda", "priced": "Narxlangan", "delivered": "Yetkazildi"}

    for i, r in enumerate(requests, 1):
        order_num = r.order.order_number if r.order else ""
        usta_name = r.usta.full_name if r.usta else ""
        mat = float(r.material_price or 0)
        dlv = float(r.delivery_price or 0)
        ext = float(r.extra_cost or 0)
        total = mat + dlv + ext
        ws.append([
            i, order_num, usta_name, float(r.amount_tonnes),
            mat, dlv, ext, total,
            STATUS_MAP.get(r.status.value, r.status.value),
            r.created_at.strftime("%d.%m.%Y %H:%M"),
        ])
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=col).fill = _ALT_FILL

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"material_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename=fname),
        caption=f"📦 Material so'rovlar — {len(requests)} ta yozuv",
    )
